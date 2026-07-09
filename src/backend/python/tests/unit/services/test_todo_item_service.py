import io
from datetime import datetime, timezone
from unittest.mock import MagicMock

import pytest
from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook

from shared.models.todo_item import TodoItem
from api.schemas.todo_item import (
    CreateTodoItemRequest,
    UpdateTodoItemRequest,
)
from api.services.todo_item_service import TodoItemService


def _make_upload_file(content: str) -> UploadFile:
    return UploadFile(filename="todo_items.csv", file=io.BytesIO(content.encode("utf-8")))


def _make_excel_upload_file(rows: list[list]) -> UploadFile:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return UploadFile(filename="todo_items.xlsx", file=buffer)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_todo(
    id: int = 1,
    title: str = "Test Todo",
    description: str | None = None,
    is_completed: bool = False,
) -> TodoItem:
    todo = TodoItem(title=title, description=description, is_completed=is_completed)
    todo.id = id
    todo.created_at = datetime(2024, 1, 1, 12, 0, 0, tzinfo=timezone.utc)
    todo.updated_at = None
    return todo


# ── get_all ───────────────────────────────────────────────────────────────────

class TestGetAll:
    def test_returns_items_and_total(self):
        repo = MagicMock()
        repo.get_all.return_value = ([_make_todo(1, "A"), _make_todo(2, "B")], 2)

        result = TodoItemService(repo).get_all(page=1, page_size=20)

        assert result.total == 2
        assert len(result.items) == 2
        assert result.items[0].title == "A"

    def test_calculates_skip_from_page(self):
        repo = MagicMock()
        repo.get_all.return_value = ([], 50)

        TodoItemService(repo).get_all(page=3, page_size=10)

        repo.get_all.assert_called_once_with(skip=20, limit=10)

    def test_total_pages_uses_ceiling_division(self):
        repo = MagicMock()
        repo.get_all.return_value = ([], 21)

        result = TodoItemService(repo).get_all(page=1, page_size=20)

        assert result.total_pages == 2

    def test_empty_result(self):
        repo = MagicMock()
        repo.get_all.return_value = ([], 0)

        result = TodoItemService(repo).get_all()

        assert result.total == 0
        assert result.items == []


# ── get_incomplete ────────────────────────────────────────────────────────────

class TestGetIncomplete:
    def test_returns_incomplete_items(self):
        repo = MagicMock()
        repo.get_incomplete.return_value = ([_make_todo(1, "Undone")], 1)

        result = TodoItemService(repo).get_incomplete(page=1, page_size=20)

        assert result.total == 1
        repo.get_incomplete.assert_called_once_with(skip=0, limit=20)

    def test_calculates_skip_from_page(self):
        repo = MagicMock()
        repo.get_incomplete.return_value = ([], 0)

        TodoItemService(repo).get_incomplete(page=2, page_size=5)

        repo.get_incomplete.assert_called_once_with(skip=5, limit=5)


# ── get_by_id ─────────────────────────────────────────────────────────────────

class TestGetById:
    def test_returns_item_when_found(self):
        repo = MagicMock()
        repo.get_by_id.return_value = _make_todo(1, "Found")

        result = TodoItemService(repo).get_by_id(1)

        assert result.id == 1
        assert result.title == "Found"

    def test_raises_404_when_not_found(self):
        repo = MagicMock()
        repo.get_by_id.return_value = None

        with pytest.raises(HTTPException) as exc_info:
            TodoItemService(repo).get_by_id(99)

        assert exc_info.value.status_code == 404
        assert "99" in exc_info.value.detail


# ── create ────────────────────────────────────────────────────────────────────

class TestCreate:
    def test_creates_and_returns_item(self):
        repo = MagicMock()
        repo.add.return_value = _make_todo(1, "New Todo", "Desc")

        result = TodoItemService(repo).create(CreateTodoItemRequest(title="New Todo", description="Desc"))

        assert result.id == 1
        assert result.title == "New Todo"
        assert result.description == "Desc"
        repo.add.assert_called_once()

    def test_creates_without_description(self):
        repo = MagicMock()
        repo.add.return_value = _make_todo(1, "No Desc")

        result = TodoItemService(repo).create(CreateTodoItemRequest(title="No Desc"))

        assert result.description is None

    def test_passes_title_and_description_to_repo(self):
        repo = MagicMock()
        repo.add.return_value = _make_todo(1, "Buy milk", "Whole milk")

        TodoItemService(repo).create(CreateTodoItemRequest(title="Buy milk", description="Whole milk"))

        added: TodoItem = repo.add.call_args[0][0]
        assert added.title == "Buy milk"
        assert added.description == "Whole milk"


# ── update ────────────────────────────────────────────────────────────────────

class TestUpdate:
    def test_updates_title(self):
        repo = MagicMock()
        todo = _make_todo(1, "Old Title")
        repo.get_by_id.return_value = todo
        repo.update.return_value = _make_todo(1, "New Title")

        result = TodoItemService(repo).update(1, UpdateTodoItemRequest(title="New Title"))

        assert result.title == "New Title"
        assert todo.title == "New Title"

    def test_updates_description(self):
        repo = MagicMock()
        todo = _make_todo(1, "Title", "Old Desc")
        repo.get_by_id.return_value = todo
        repo.update.return_value = todo

        TodoItemService(repo).update(1, UpdateTodoItemRequest(description="New Desc"))

        assert todo.description == "New Desc"

    def test_updates_is_completed(self):
        repo = MagicMock()
        todo = _make_todo(1, "Title", is_completed=False)
        repo.get_by_id.return_value = todo
        repo.update.return_value = todo

        TodoItemService(repo).update(1, UpdateTodoItemRequest(is_completed=True))

        assert todo.is_completed is True

    def test_skips_none_fields(self):
        repo = MagicMock()
        todo = _make_todo(1, "Original", "Original Desc", False)
        repo.get_by_id.return_value = todo
        repo.update.return_value = todo

        TodoItemService(repo).update(1, UpdateTodoItemRequest())

        assert todo.title == "Original"
        assert todo.description == "Original Desc"
        assert todo.is_completed is False

    def test_raises_404_when_not_found(self):
        repo = MagicMock()
        repo.get_by_id.return_value = None

        with pytest.raises(HTTPException) as exc_info:
            TodoItemService(repo).update(99, UpdateTodoItemRequest(title="X"))

        assert exc_info.value.status_code == 404


# ── delete ────────────────────────────────────────────────────────────────────

class TestDelete:
    def test_deletes_item(self):
        repo = MagicMock()
        todo = _make_todo(1)
        repo.get_by_id.return_value = todo

        TodoItemService(repo).delete(1)

        repo.delete.assert_called_once_with(todo)

    def test_raises_404_when_not_found(self):
        repo = MagicMock()
        repo.get_by_id.return_value = None

        with pytest.raises(HTTPException) as exc_info:
            TodoItemService(repo).delete(99)

        assert exc_info.value.status_code == 404


# ── mark_complete ─────────────────────────────────────────────────────────────

class TestMarkComplete:
    def test_sets_is_completed_to_true(self):
        repo = MagicMock()
        todo = _make_todo(1, "Pending", is_completed=False)
        repo.get_by_id.return_value = todo
        repo.update.return_value = _make_todo(1, "Pending", is_completed=True)

        result = TodoItemService(repo).mark_complete(1)

        assert todo.is_completed is True
        assert result.is_completed is True
        repo.update.assert_called_once_with(todo)

    def test_raises_404_when_not_found(self):
        repo = MagicMock()
        repo.get_by_id.return_value = None

        with pytest.raises(HTTPException) as exc_info:
            TodoItemService(repo).mark_complete(99)

        assert exc_info.value.status_code == 404


# ── import_csv ────────────────────────────────────────────────────────────────

class TestImportCsv:
    def test_imports_valid_rows(self):
        repo = MagicMock()
        csv_content = "title,description,is_completed\nBuy milk,Whole milk,false\nWalk dog,,true\n"

        result = TodoItemService(repo).import_csv(_make_upload_file(csv_content))

        assert result.imported == 2
        assert result.failed == 0
        assert result.errors == []
        assert repo.add.call_count == 2

    def test_maps_row_fields_onto_todo_item(self):
        repo = MagicMock()
        csv_content = "title,description,is_completed\nBuy milk,Whole milk,true\n"

        TodoItemService(repo).import_csv(_make_upload_file(csv_content))

        added: TodoItem = repo.add.call_args[0][0]
        assert added.title == "Buy milk"
        assert added.description == "Whole milk"
        assert added.is_completed is True

    def test_reports_error_for_missing_title(self):
        repo = MagicMock()
        csv_content = "title,description,is_completed\n,No title,false\nValid,ok,false\n"

        result = TodoItemService(repo).import_csv(_make_upload_file(csv_content))

        assert result.imported == 1
        assert result.failed == 1
        assert result.errors[0].row == 2
        assert "Title is required" in result.errors[0].error

    def test_defaults_missing_description_and_is_completed(self):
        repo = MagicMock()
        csv_content = "title\nJust a title\n"

        result = TodoItemService(repo).import_csv(_make_upload_file(csv_content))

        assert result.imported == 1
        added: TodoItem = repo.add.call_args[0][0]
        assert added.description is None
        assert added.is_completed is False

    def test_empty_file_imports_nothing(self):
        repo = MagicMock()
        csv_content = "title,description,is_completed\n"

        result = TodoItemService(repo).import_csv(_make_upload_file(csv_content))

        assert result.imported == 0
        assert result.failed == 0
        repo.add.assert_not_called()


# ── export_csv ────────────────────────────────────────────────────────────────

class TestExportCsv:
    def test_returns_header_and_rows(self):
        repo = MagicMock()
        repo.get_all_items.return_value = [_make_todo(1, "A", "Desc A", True)]

        content = TodoItemService(repo).export_csv()

        lines = content.strip().splitlines()
        assert lines[0] == "id,title,description,is_completed,created_at,updated_at"
        assert lines[1].startswith("1,A,Desc A,True,")

    def test_empty_list_returns_header_only(self):
        repo = MagicMock()
        repo.get_all_items.return_value = []

        content = TodoItemService(repo).export_csv()

        lines = content.strip().splitlines()
        assert len(lines) == 1
        assert lines[0] == "id,title,description,is_completed,created_at,updated_at"

    def test_handles_null_description(self):
        repo = MagicMock()
        repo.get_all_items.return_value = [_make_todo(1, "A", None, False)]

        content = TodoItemService(repo).export_csv()

        lines = content.strip().splitlines()
        assert lines[1] == "1,A,,False,2024-01-01T12:00:00+00:00,"


# ── import_excel ──────────────────────────────────────────────────────────────

class TestImportExcel:
    def test_imports_valid_rows(self):
        repo = MagicMock()
        file = _make_excel_upload_file([
            ["title", "description", "is_completed"],
            ["Buy milk", "Whole milk", False],
            ["Walk dog", None, True],
        ])

        result = TodoItemService(repo).import_excel(file)

        assert result.imported == 2
        assert result.failed == 0
        assert result.errors == []
        assert repo.add.call_count == 2

    def test_maps_row_fields_onto_todo_item(self):
        repo = MagicMock()
        file = _make_excel_upload_file([
            ["title", "description", "is_completed"],
            ["Buy milk", "Whole milk", True],
        ])

        TodoItemService(repo).import_excel(file)

        added: TodoItem = repo.add.call_args[0][0]
        assert added.title == "Buy milk"
        assert added.description == "Whole milk"
        assert added.is_completed is True

    def test_reports_error_for_missing_title(self):
        repo = MagicMock()
        file = _make_excel_upload_file([
            ["title", "description", "is_completed"],
            [None, "No title", False],
            ["Valid", "ok", False],
        ])

        result = TodoItemService(repo).import_excel(file)

        assert result.imported == 1
        assert result.failed == 1
        assert result.errors[0].row == 2
        assert "Title is required" in result.errors[0].error

    def test_defaults_missing_description_and_is_completed(self):
        repo = MagicMock()
        file = _make_excel_upload_file([
            ["title"],
            ["Just a title"],
        ])

        result = TodoItemService(repo).import_excel(file)

        assert result.imported == 1
        added: TodoItem = repo.add.call_args[0][0]
        assert added.description is None
        assert added.is_completed is False

    def test_empty_file_imports_nothing(self):
        repo = MagicMock()
        file = _make_excel_upload_file([
            ["title", "description", "is_completed"],
        ])

        result = TodoItemService(repo).import_excel(file)

        assert result.imported == 0
        assert result.failed == 0
        repo.add.assert_not_called()


# ── export_excel ──────────────────────────────────────────────────────────────

class TestExportExcel:
    def test_returns_header_and_rows(self):
        repo = MagicMock()
        repo.get_all_items.return_value = [_make_todo(1, "A", "Desc A", True)]

        content = TodoItemService(repo).export_excel()

        sheet = load_workbook(io.BytesIO(content)).active
        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0] == ("id", "title", "description", "is_completed", "created_at", "updated_at")
        assert rows[1][0] == 1
        assert rows[1][1] == "A"
        assert rows[1][2] == "Desc A"
        assert rows[1][3] is True

    def test_empty_list_returns_header_only(self):
        repo = MagicMock()
        repo.get_all_items.return_value = []

        content = TodoItemService(repo).export_excel()

        sheet = load_workbook(io.BytesIO(content)).active
        rows = list(sheet.iter_rows(values_only=True))
        assert len(rows) == 1
        assert rows[0] == ("id", "title", "description", "is_completed", "created_at", "updated_at")

    def test_handles_null_description(self):
        repo = MagicMock()
        repo.get_all_items.return_value = [_make_todo(1, "A", None, False)]

        content = TodoItemService(repo).export_excel()

        sheet = load_workbook(io.BytesIO(content)).active
        rows = list(sheet.iter_rows(values_only=True))
        assert rows[1] == (1, "A", None, False, "2024-01-01T12:00:00+00:00", None)
