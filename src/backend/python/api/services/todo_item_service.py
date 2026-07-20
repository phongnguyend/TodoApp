import csv
import io
from abc import ABC, abstractmethod

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from shared.models.todo_item import TodoItem
from api.repositories.todo_item_repository import ITodoItemRepository, TodoItemRepository
from api.schemas.todo_item import (
    CreateTodoItemRequest,
    ImportResult,
    ImportRowError,
    PaginatedResponse,
    TodoItemResponse,
    UpdateTodoItemRequest,
)

# ── CSV import/export helpers ───────────────────────────────────────────────────

_CSV_FIELDNAMES = ["id", "title", "description", "is_completed", "created_at", "updated_at"]
_TRUE_VALUES = {"1", "true", "yes", "y"}


def _parse_bool(value: str | None) -> bool:
    return (value or "").strip().lower() in _TRUE_VALUES


def _parse_bool_cell(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return _parse_bool(str(value) if value is not None else None)


class ITodoItemService(ABC):
    """Service interface - mirrors ITodoService in C#."""

    @abstractmethod
    def get_all(self, page: int, page_size: int) -> PaginatedResponse[TodoItemResponse]: ...

    @abstractmethod
    def get_incomplete(self, page: int, page_size: int) -> PaginatedResponse[TodoItemResponse]: ...

    @abstractmethod
    def get_by_id(self, todo_id: int) -> TodoItemResponse: ...

    @abstractmethod
    def create(self, request: CreateTodoItemRequest, actor_user_id: int | None = None) -> TodoItemResponse: ...

    @abstractmethod
    def update(self, todo_id: int, request: UpdateTodoItemRequest, actor_user_id: int | None = None) -> TodoItemResponse: ...

    @abstractmethod
    def delete(self, todo_id: int) -> None: ...

    @abstractmethod
    def mark_complete(self, todo_id: int, actor_user_id: int | None = None) -> TodoItemResponse: ...

    @abstractmethod
    def import_csv(self, file: UploadFile, actor_user_id: int | None = None) -> ImportResult: ...

    @abstractmethod
    def export_csv(self) -> str: ...

    @abstractmethod
    def import_excel(self, file: UploadFile, actor_user_id: int | None = None) -> ImportResult: ...

    @abstractmethod
    def export_excel(self) -> bytes: ...


class TodoItemService(ITodoItemService):
    """Business-logic layer (analogous to an ASP.NET Core service registered via DI)."""

    def __init__(self, repository: ITodoItemRepository) -> None:
        self._repo = repository

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _get_or_404(self, todo_id: int) -> TodoItem:
        todo = self._repo.get_by_id(todo_id)
        if todo is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Todo item {todo_id} not found.")
        return todo

    @staticmethod
    def _to_paginated(items: list[TodoItem], total: int, page: int, page_size: int) -> PaginatedResponse[TodoItemResponse]:
        return PaginatedResponse(
            items=[TodoItemResponse.model_validate(item) for item in items],
            total=total,
            page=page,
            page_size=page_size,
            total_pages=-(-total // page_size),  # ceiling division
        )

    # ── Queries ───────────────────────────────────────────────────────────────

    def get_all(self, page: int = 1, page_size: int = 20) -> PaginatedResponse[TodoItemResponse]:
        skip = (page - 1) * page_size
        items, total = self._repo.get_all(skip=skip, limit=page_size)
        return self._to_paginated(items, total, page, page_size)

    def get_incomplete(self, page: int = 1, page_size: int = 20) -> PaginatedResponse[TodoItemResponse]:
        skip = (page - 1) * page_size
        items, total = self._repo.get_incomplete(skip=skip, limit=page_size)
        return self._to_paginated(items, total, page, page_size)

    def get_by_id(self, todo_id: int) -> TodoItemResponse:
        todo = self._get_or_404(todo_id)
        return TodoItemResponse.model_validate(todo)

    # ── Commands ──────────────────────────────────────────────────────────────

    def create(self, request: CreateTodoItemRequest, actor_user_id: int | None = None) -> TodoItemResponse:
        todo = TodoItem(title=request.title, description=request.description, created_by_user_id=actor_user_id)
        created = self._repo.add(todo)
        return TodoItemResponse.model_validate(created)

    def update(self, todo_id: int, request: UpdateTodoItemRequest, actor_user_id: int | None = None) -> TodoItemResponse:
        todo = self._get_or_404(todo_id)
        if request.title is not None:
            todo.title = request.title
        if request.description is not None:
            todo.description = request.description
        if request.is_completed is not None:
            todo.is_completed = request.is_completed
        todo.updated_by_user_id = actor_user_id
        updated = self._repo.update(todo)
        return TodoItemResponse.model_validate(updated)

    def delete(self, todo_id: int) -> None:
        todo = self._get_or_404(todo_id)
        self._repo.delete(todo)

    def mark_complete(self, todo_id: int, actor_user_id: int | None = None) -> TodoItemResponse:
        todo = self._get_or_404(todo_id)
        todo.is_completed = True
        todo.updated_by_user_id = actor_user_id
        updated = self._repo.update(todo)
        return TodoItemResponse.model_validate(updated)

    # ── CSV import/export ─────────────────────────────────────────────────────

    def import_csv(self, file: UploadFile, actor_user_id: int | None = None) -> ImportResult:
        raw = file.file.read()
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))

        imported = 0
        errors: list[ImportRowError] = []
        for row_number, row in enumerate(reader, start=2):  # header is row 1
            title = (row.get("title") or "").strip()
            if not title:
                errors.append(ImportRowError(row=row_number, error="Title is required."))
                continue

            description = (row.get("description") or "").strip() or None
            is_completed = _parse_bool(row.get("is_completed"))

            todo = TodoItem(title=title, description=description, is_completed=is_completed,
                            created_by_user_id=actor_user_id)
            self._repo.add(todo)
            imported += 1

        return ImportResult(imported=imported, failed=len(errors), errors=errors)

    def export_csv(self) -> str:
        items = self._repo.get_all_items()

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(_CSV_FIELDNAMES)
        for item in items:
            writer.writerow([
                item.id,
                item.title,
                item.description or "",
                item.is_completed,
                item.created_at.isoformat() if item.created_at else "",
                item.updated_at.isoformat() if item.updated_at else "",
            ])
        return buffer.getvalue()

    # ── Excel import/export ────────────────────────────────────────────────────

    def import_excel(self, file: UploadFile, actor_user_id: int | None = None) -> ImportResult:
        raw = file.file.read()
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active

        rows = sheet.iter_rows(values_only=True)
        header = [str(cell).strip().lower() if cell is not None else "" for cell in next(rows, ())]
        col_index = {name: idx for idx, name in enumerate(header)}

        def _cell(row: tuple, name: str) -> object:
            idx = col_index.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        imported = 0
        errors: list[ImportRowError] = []
        for row_number, row in enumerate(rows, start=2):  # header is row 1
            if row is None or all(value is None for value in row):
                continue

            title = str(_cell(row, "title") or "").strip()
            if not title:
                errors.append(ImportRowError(row=row_number, error="Title is required."))
                continue

            description = str(_cell(row, "description") or "").strip() or None
            is_completed = _parse_bool_cell(_cell(row, "is_completed"))

            todo = TodoItem(title=title, description=description, is_completed=is_completed,
                            created_by_user_id=actor_user_id)
            self._repo.add(todo)
            imported += 1

        return ImportResult(imported=imported, failed=len(errors), errors=errors)

    def export_excel(self) -> bytes:
        items = self._repo.get_all_items()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Todo Items"
        sheet.append(_CSV_FIELDNAMES)
        for item in items:
            sheet.append([
                item.id,
                item.title,
                item.description or "",
                item.is_completed,
                item.created_at.isoformat() if item.created_at else "",
                item.updated_at.isoformat() if item.updated_at else "",
            ])

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()


# ── Dependency factory (used by FastAPI Depends) ──────────────────────────────

def get_todo_service(db: Session) -> ITodoItemService:
    repository = TodoItemRepository(db)
    return TodoItemService(repository)
